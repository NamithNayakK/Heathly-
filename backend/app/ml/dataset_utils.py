from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


PHQ9_COLUMNS = [
    "Interest",
    "Depressed",
    "Sleep",
    "Energy",
    "Appetite",
    "Self Worth",
    "Concentration",
    "Movement",
    "Self Harm Thoughts",
]

BASE_CONTEXT_COLUMNS = [
    "Campus Environment",
    "Approach Professors",
    "Academic Support",
    "Peer Support",
    "Peer Pressure",
    "Physical Activity Type",
    "Activity Time",
    "Age Group",
    "Gender",
]


def _normalize_text(value: object) -> str:
    return str(value or "").strip().lower()


def _response_to_score(value: object) -> int:
    normalized = _normalize_text(value)
    mapping = {
        "not at all": 0,
        "several days": 1,
        "more than half the days": 2,
        "nearly every day": 3,
    }
    return mapping.get(normalized, 0)


@dataclass(frozen=True)
class SurveySample:
    answers: list[int]
    text: str
    emotion_label: str
    mental_state_label: str


def resolve_dataset_path(dataset_path: str | Path | None = None) -> Path:
    if dataset_path is not None:
        return Path(dataset_path)
    return Path(__file__).resolve().parents[3] / "mental_wellness_dataset_u.xlsx"


def load_survey_samples(dataset_path: str | Path | None = None) -> list[SurveySample]:
    path = resolve_dataset_path(dataset_path)
    if not path.exists():
        raise FileNotFoundError(f"Dataset not found at {path}")

    workbook = load_workbook(path, read_only=False, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [cell.value for cell in sheet[1]]
    header_index = {header: position + 1 for position, header in enumerate(headers) if header}

    samples: list[SurveySample] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_values = list(row)
        answers = [_response_to_score(row_values[header_index[column] - 1]) for column in PHQ9_COLUMNS]
        context = {name: row_values[header_index[name] - 1] for name in BASE_CONTEXT_COLUMNS if name in header_index}
        text = build_emotion_text(answers, context)
        samples.append(
            SurveySample(
                answers=answers,
                text=text,
                emotion_label=infer_emotion_label(answers),
                mental_state_label=infer_mental_state_label(answers),
            )
        )

    return samples


def build_emotion_text(answers: list[int], context: dict[str, object] | None = None) -> str:
    context = context or {}
    item_text = [
        f"interest is {describe_level(answers[0])}",
        f"mood is {describe_level(answers[1])}",
        f"sleep is {describe_level(answers[2])}",
        f"energy is {describe_level(answers[3])}",
        f"appetite is {describe_level(answers[4])}",
        f"self worth is {describe_level(answers[5])}",
        f"concentration is {describe_level(answers[6])}",
        f"movement is {describe_level(answers[7])}",
        f"self harm thoughts are {describe_level(answers[8])}",
    ]

    context_bits = []
    for key in ["Campus Environment", "Academic Support", "Peer Support", "Peer Pressure", "Physical Activity Type", "Activity Time"]:
        value = context.get(key)
        if value:
            context_bits.append(f"{key.lower()} is {value}")

    return "Student reports " + ", ".join(item_text) + ". " + ". ".join(context_bits or ["no additional context"])


def describe_level(score: int) -> str:
    mapping = {
        0: "not at all",
        1: "several days",
        2: "more than half the days",
        3: "nearly every day",
    }
    return mapping.get(int(score), "not at all")


def infer_emotion_label(answers: list[int]) -> str:
    total = sum(answers)
    self_harm = answers[8]
    depressed = answers[1]
    sleep = answers[2]
    energy = answers[3]
    self_worth = answers[5]
    concentration = answers[6]
    movement = answers[7]
    interest = answers[0]

    if self_harm >= 2 or (depressed >= 2 and self_worth >= 2):
        return "sadness"
    if sleep >= 2 or concentration >= 2:
        return "anxiety"
    if self_worth >= 2:
        return "shame"
    if energy >= 2 or movement >= 2 or total >= 15:
        return "overwhelm"
    if total <= 4 and interest <= 1 and depressed <= 1:
        return "stable"
    return "sadness"


def infer_mental_state_label(answers: list[int]) -> str:
    total = sum(answers)
    self_harm = answers[8]

    if self_harm >= 2:
        return "crisis"
    if total >= 20:
        return "severe_distress"
    if total >= 15:
        return "moderate_distress"
    if total >= 8:
        return "mild_distress"
    return "stable"


def split_samples(samples: list[SurveySample], validation_ratio: float = 0.2, seed: int = 42) -> tuple[list[SurveySample], list[SurveySample]]:
    import random

    rng = random.Random(seed)
    shuffled = samples[:]
    rng.shuffle(shuffled)
    split_index = max(1, int(len(shuffled) * (1 - validation_ratio)))
    return shuffled[:split_index], shuffled[split_index:]


def label_lookup(labels: Iterable[str]) -> tuple[dict[str, int], dict[int, str]]:
    unique = sorted(set(labels))
    label2id = {label: index for index, label in enumerate(unique)}
    id2label = {index: label for label, index in label2id.items()}
    return label2id, id2label
